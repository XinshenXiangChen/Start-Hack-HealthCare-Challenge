#!/usr/bin/env python3
"""Standardize mixed source files into DB target-table CSVs."""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import subprocess
import unicodedata
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pdfplumber


CREATE_RE = re.compile(r"^\s*create\s+table\s+([A-Za-z0-9_]+)\s*$", re.IGNORECASE)
EPA_CODE_RE = re.compile(r"E\d+_I_\d+", re.IGNORECASE)
AC_LEGACY_EPA_RE = re.compile(r"^EPA(\d{3,6})[A-Z0-9_]*$", re.IGNORECASE)
JSON_BLOCK_RE = re.compile(r"\{.*\}", re.DOTALL)
PDF_DATE_RE = re.compile(r"Date:\s*(.*?)\s+Shift:", re.IGNORECASE | re.DOTALL)
PDF_SHIFT_RE = re.compile(
    r"Shift:\s*(.*?)\s+Patient Information", re.IGNORECASE | re.DOTALL
)
PDF_PAT_RE = re.compile(r"Patient ID:\s*(.*?)\s+Case ID:", re.IGNORECASE | re.DOTALL)
PDF_CASE_RE = re.compile(r"Case ID:\s*(.*?)\s+Ward:", re.IGNORECASE | re.DOTALL)
PDF_WARD_RE = re.compile(r"Ward:\s*(.*?)\s+Report", re.IGNORECASE | re.DOTALL)
SQL_INSERT_RE = re.compile(
    r"insert\s+into\s+([^\s(]+)\s*(?:\((.*?)\))?\s*values\s*(.*?);",
    re.IGNORECASE | re.DOTALL,
)
SID_CODE_RE = re.compile(r"\b\d{2}(?:[_-]\d{2})+\b")
IID_DICT_FILENAMES = ("IID-SID-ITEM.csv", "iid-sid-item.csv", "IID_SID_ITEM.csv")

_IID_LOOKUP: dict[str, str] = {}
_IID_DICTIONARY_SOURCE: str | None = None
_IID_LOOKUP_INITIALIZED = False


def parse_bool_env(name: str, default: bool = False) -> bool:
    text = (os.getenv(name) or "").strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "on"}


DEBUG_ROUTING = parse_bool_env("PIPELINE_DEBUG_ROUTING", False)


def log_routing_debug(payload: dict[str, Any]) -> None:
    if not DEBUG_ROUTING:
        return
    print(
        "[routing-debug] " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        flush=True,
    )


TABLE_HINTS = {
    "tbImportLabsData": ["lab", "labs"],
    "tbImportIcd10Data": ["icd", "ops"],
    "tbImportDevice1HzMotionData": ["1hz", "raw_1hz"],
    "tbImportDeviceMotionData": ["device", "motion", "fall"],
    "tbImportMedicationInpatientData": ["medication", "med"],
    "tbImportNursingDailyReportsData": ["nursing", "report"],
    "tbImportAcData": ["epaac", "epaac-data", "einsch"],
}


AC_LEGACY_CASE_KEYS = {"patfal", "fallid", "fall_id", "caseid", "case_id", "cas"}
AC_LEGACY_CASE_ALPHA_KEYS = {"patfoe", "patdoe"}
AC_SID_COL_KEYS = {"sid", "itmsid", "itemsid", "sidcode", "itemcode"}
AC_SID_VALUE_COL_KEYS = {
    "sid_value",
    "sidvalue",
    "itemvalue",
    "value",
    "val",
    "wert",
}
ENTITY_ID_COLS = {"coCaseId", "coCaseIdAlpha", "coPatient_id", "coEncounter_id"}
NULL_LIKE_TEXT = {"na", "n/a", "nan", "none", "null", "unknown", "-"}


ALIASES = {
    "tbImportLabsData": {
        "caseid": "coCaseId",
        "case_id": "coCaseId",
        "cas": "coCaseId",
        "specdt": "coSpecimen_datetime",
        "specimen_datetime": "coSpecimen_datetime",
        "na": "coSodium_mmol_L",
        "na_flag": "coSodium_flag",
        "na_low": "cosodium_ref_low",
        "na_high": "cosodium_ref_high",
        "k": "coPotassium_mmol_L",
        "k_flag": "coPotassium_flag",
        "k_low": "coPotassium_ref_low",
        "k_high": "coPotassium_ref_high",
        "creat": "coCreatinine_mg_dL",
        "creat_flag": "coCreatinine_flag",
        "creat_low": "coCreatinine_ref_low",
        "creat_high": "coCreatinine_ref_high",
        "egfr": "coEgfr_mL_min_1_73m2",
        "egfr_flag": "coEgfr_flag",
        "egfr_low": "coEgfr_ref_low",
        "egfr_high": "coEgfr_ref_high",
        "gluc": "coGlucose_mg_dL",
        "gluc_flag": "coGlucose_flag",
        "gluc_low": "coGlucose_ref_low",
        "gluc_high": "coGlucose_ref_high",
        "hb": "coHemoglobin_g_dL",
        "hb_flag": "coHb_flag",
        "hb_low": "coHb_ref_low",
        "hb_high": "coHb_ref_high",
        "wbc": "coWbc_10e9_L",
        "wbc_flag": "coWbc_flag",
        "wbc_low": "coWbc_ref_low",
        "wbc_high": "coWbc_ref_high",
        "plt": "coPlatelets_10e9_L",
        "plt_flag": "coPlatelets_flag",
        "plt_low": "coPlt_ref_low",
        "plt_high": "coPlt_ref_high",
        "crp": "coCrp_mg_L",
        "crp_flag": "coCrp_flag",
        "crp_low": "coCrp_ref_low",
        "crp_high": "coCrp_ref_high",
        "alt": "coAlt_U_L",
        "alt_flag": "coAlt_flag",
        "alt_low": "coAlt_ref_low",
        "alt_high": "coAlt_ref_high",
        "ast": "coAst_U_L",
        "ast_flag": "coAst_flag",
        "ast_low": "coAst_ref_low",
        "ast_high": "coAst_ref_high",
        "bili": "coBilirubin_mg_dL",
        "bili_flag": "coBilirubin_flag",
        "bili_low": "coBili_ref_low",
        "bili_high": "coBili_ref_high",
        "alb": "coAlbumin_g_dL",
        "alb_flag": "coAlbumin_flag",
        "alb_low": "coAlbumin_ref_low",
        "alb_high": "coAlbumin_ref_high",
        "inr": "coInr",
        "inr_flag": "coInr_flag",
        "inr_low": "coInr_ref_low",
        "inr_high": "coInr_ref_high",
        "lact": "coLactate_mmol_L",
        "lact_flag": "coLactate_flag",
        "lact_low": "coLactate_ref_low",
        "lact_high": "coLactate_ref_high",
    },
    "tbImportIcd10Data": {
        "caseid": "coCaseId",
        "case_id": "coCaseId",
        "station": "coWard",
        "ward": "coWard",
        "aufnahmedatum": "coAdmission_date",
        "admission_date": "coAdmission_date",
        "entlassungsdatum": "coDischarge_date",
        "discharge_date": "coDischarge_date",
        "verweildauer_tage": "coLength_of_stay_days",
        "length_of_stay_days": "coLength_of_stay_days",
        "icd10_haupt": "coPrimary_icd10_code",
        "primary_icd10_code": "coPrimary_icd10_code",
        "icd10_haupt_bezeichnung": "coPrimary_icd10_description_en",
        "primary_icd10_description_en": "coPrimary_icd10_description_en",
        "icd10_neben": "coSecondary_icd10_codes",
        "secondary_icd10_codes": "coSecondary_icd10_codes",
        "icd10_neben_bezeichnung": "cpSecondary_icd10_descriptions_en",
        "secondary_icd10_descriptions_en": "cpSecondary_icd10_descriptions_en",
        "ops_code": "coOps_codes",
        "ops_codes": "coOps_codes",
        "ops_bezeichnung": "ops_descriptions_en",
        "ops_descriptions_en": "ops_descriptions_en",
        "id_cas": "coCaseId",
        "date_ad": "coAdmission_date",
        "d_s": "coPrimary_icd10_code",
        "d_s_str": "coPrimary_icd10_description_en",
        "d_m": "coSecondary_icd10_codes",
        "d_m_str": "cpSecondary_icd10_descriptions_en",
        "proc": "coOps_codes",
        "proc_str": "ops_descriptions_en",
    },
    "tbImportDeviceMotionData": {
        "timestamp": "coTimestamp",
        "date": "coTimestamp",
        "patient_id": "coPatient_id",
        "id": "coPatient_id",
        "movement_index_0_100": "coMovement_index_0_100",
        "idx_mov": "coMovement_index_0_100",
        "micro_movements_count": "coMicro_movements_count",
        "num_mov": "coMicro_movements_count",
        "bed_exit_detected_0_1": "coBed_exit_detected_0_1",
        "num_ex": "coBed_exit_detected_0_1",
        "fall_event_0_1": "coFall_event_0_1",
        "fall": "coFall_event_0_1",
        "impact_magnitude_g": "coImpact_magnitude_g",
        "mag_impact": "coImpact_magnitude_g",
        "post_fall_immobility_minutes": "coPost_fall_immobility_minutes",
        "min_immob": "coPost_fall_immobility_minutes",
    },
    "tbImportDevice1HzMotionData": {
        "timestamp": "coTimestamp",
        "date": "coTimestamp",
        "patient_id": "coPatient_id",
        "id_pat": "coPatient_id",
        "patientid": "coPatient_id",
        "device_id": "coDevice_id",
        "id_dev": "coDevice_id",
        "deviceid": "coDevice_id",
        "bed_occupied_0_1": "coBed_occupied_0_1",
        "bedoccupied": "coBed_occupied_0_1",
        "movement_score_0_100": "coMovement_score_0_100",
        "movementscore": "coMovement_score_0_100",
        "accel_x_m_s2": "coAccel_x_m_s2",
        "accelx": "coAccel_x_m_s2",
        "accx": "coAccel_x_m_s2",
        "accel_y_m_s2": "coAccel_y_m_s2",
        "accely": "coAccel_y_m_s2",
        "accy": "coAccel_y_m_s2",
        "accel_z_m_s2": "coAccel_z_m_s2",
        "accelz": "coAccel_z_m_s2",
        "accz": "coAccel_z_m_s2",
        "accel_magnitude_g": "coAccel_magnitude_g",
        "accelmag": "coAccel_magnitude_g",
        "acc_mag": "coAccel_magnitude_g",
        "pressure_zone1_0_100": "coPressure_zone1_0_100",
        "pressz1": "coPressure_zone1_0_100",
        "p1": "coPressure_zone1_0_100",
        "pressure_zone2_0_100": "coPressure_zone2_0_100",
        "pressz2": "coPressure_zone2_0_100",
        "p2": "coPressure_zone2_0_100",
        "pressure_zone3_0_100": "coPressure_zone3_0_100",
        "pressz3": "coPressure_zone3_0_100",
        "p3": "coPressure_zone3_0_100",
        "pressure_zone4_0_100": "coPressure_zone4_0_100",
        "pressz4": "coPressure_zone4_0_100",
        "p4": "coPressure_zone4_0_100",
        "bed_exit_event_0_1": "coBed_exit_event_0_1",
        "bedexit": "coBed_exit_event_0_1",
        "exit": "coBed_exit_event_0_1",
        "bed_return_event_0_1": "coBed_return_event_0_1",
        "bedreturn": "coBed_return_event_0_1",
        "return": "coBed_return_event_0_1",
        "fall_event_0_1": "coFall_event_0_1",
        "fallevent": "coFall_event_0_1",
        "fall": "coFall_event_0_1",
        "impact_magnitude_g": "coImpact_magnitude_g",
        "impactmag": "coImpact_magnitude_g",
        "event_id": "coEvent_id",
        "eventid": "coEvent_id",
    },
    "tbImportMedicationInpatientData": {
        "record_type": "coRecord_type",
        "rec_type": "coRecord_type",
        "patient_id": "coPatient_id",
        "pat_id": "coPatient_id",
        "encounter_id": "coEncounter_id",
        "enc_id": "coEncounter_id",
        "ward": "coWard",
        "station": "coWard",
        "admission_datetime": "coAdmission_datetime",
        "aufnahme_dt": "coAdmission_datetime",
        "discharge_datetime": "coDischarge_datetime",
        "entlassung_dt": "coDischarge_datetime",
        "order_id": "coOrder_id",
        "order_uuid": "coOrder_uuid",
        "uuid": "coOrder_uuid",
        "medication_code_atc": "coMedication_code_atc",
        "atc_code": "coMedication_code_atc",
        "medication_name": "coMedication_name",
        "medikament": "coMedication_name",
        "route": "coRoute",
        "applikation": "coRoute",
        "dose": "coDose",
        "dosis": "coDose",
        "dose_unit": "coDose_unit",
        "einheit": "coDose_unit",
        "frequency": "coFrequency",
        "haeufigkeit": "coFrequency",
        "order_start_datetime": "coOrder_start_datetime",
        "start_dt": "coOrder_start_datetime",
        "order_stop_datetime": "coOrder_stop_datetime",
        "stop_dt": "coOrder_stop_datetime",
        "is_prn_0_1": "coIs_prn_0_1",
        "bei_bedarf": "coIs_prn_0_1",
        "indication": "coIndication",
        "indikation": "coIndication",
        "prescriber_role": "prescriber_role",
        "verschreiber": "prescriber_role",
        "order_status": "order_status",
        "bestellung_status": "order_status",
        "administration_datetime": "administration_datetime",
        "gabe_dt": "administration_datetime",
        "administered_dose": "administered_dose",
        "gegebene_dosis": "administered_dose",
        "administered_unit": "administered_unit",
        "gabe_einheit": "administered_unit",
        "administration_status": "administration_status",
        "gabe_status": "administration_status",
        "note": "note",
        "notiz": "note",
    },
    "tbImportNursingDailyReportsData": {
        "caseid": "coCaseId",
        "case_id": "coCaseId",
        "cas": "coCaseId",
        "patient_id": "coPatient_id",
        "patientid": "coPatient_id",
        "pat": "coPatient_id",
        "ward": "coWard",
        "war": "coWard",
        "report_date": "coReport_date",
        "reportdate": "coReport_date",
        "dat": "coReport_date",
        "shift": "coShift",
        "shf": "coShift",
        "nursing_note_free_text": "coNursing_note_free_text",
        "nursingnote": "coNursing_note_free_text",
        "txt": "coNursing_note_free_text",
    },
}


def normalize(name: str) -> str:
    text = unicodedata.normalize("NFKD", name or "")
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9_]", "", text.strip().lower())


def to_target_col_from_iid(iid: str) -> str:
    cleaned = iid.replace("-", "_").upper().strip()
    m = re.match(r"^(E\d+)_I_(\d+)$", cleaned)
    if m:
        left = m.group(1).replace("_", "")
        num = str(int(m.group(2)))
        if len(num) < 3:
            num = num.zfill(3)
        return f"co{left}I{num}"
    return "co" + cleaned.replace("_", "")


def build_iid_lookup_keys(text: str) -> list[str]:
    raw = (text or "").strip()
    if not raw:
        return []
    out: list[str] = []
    for cand in {
        raw,
        raw.lower(),
        raw.replace("-", "_"),
        raw.lower().replace("-", "_"),
        raw.replace("_", ""),
        raw.lower().replace("_", ""),
        raw.replace("-", ""),
        raw.lower().replace("-", ""),
        normalize(raw),
    }:
        c = cand.strip()
        if c:
            out.append(c)
    return out


def parse_iid_sid_item_dictionary(path: Path) -> dict[str, str]:
    enc, delim = sniff_text_file(path)
    with path.open(encoding=enc, newline="") as f:
        reader = csv.DictReader(f, delimiter=delim)
        if not reader.fieldnames:
            return {}
        headers = {normalize(h): h for h in reader.fieldnames}

        iid_col = headers.get("itmiid") or headers.get("iid")
        sid_col = headers.get("itmsid") or headers.get("sid")
        de_col = (
            headers.get("itmname255_de")
            or headers.get("itmname_de")
            or headers.get("name_de")
            or headers.get("itemname_de")
        )
        en_col = (
            headers.get("itmname255_en")
            or headers.get("itmname_en")
            or headers.get("name_en")
            or headers.get("itemname_en")
        )

        if not iid_col:
            return {}

        lookup: dict[str, str] = {}
        for row in reader:
            iid = str(row.get(iid_col, "") or "").strip()
            if not iid:
                continue
            iid = iid.replace("-", "_").upper()
            for key in build_iid_lookup_keys(iid):
                lookup[key] = iid
            if sid_col:
                sid_val = str(row.get(sid_col, "") or "").strip()
                for key in build_iid_lookup_keys(sid_val):
                    lookup[key] = iid
            if de_col:
                de_val = str(row.get(de_col, "") or "").strip()
                for key in build_iid_lookup_keys(de_val):
                    lookup[key] = iid
            if en_col:
                en_val = str(row.get(en_col, "") or "").strip()
                for key in build_iid_lookup_keys(en_val):
                    lookup[key] = iid
        return lookup


def discover_iid_dictionary_path(
    explicit_path: Path | None, search_roots: list[Path] | None = None
) -> Path | None:
    if explicit_path and explicit_path.exists():
        return explicit_path

    env_text = (os.getenv("PIPELINE_IID_DICTIONARY") or "").strip()
    env_path = Path(env_text) if env_text else None
    if env_path and env_path.exists():
        return env_path

    roots = search_roots[:] if search_roots else []
    roots.extend([Path.cwd(), Path(__file__).resolve().parents[1]])
    seen: set[str] = set()
    for root in roots:
        root_r = root.resolve()
        root_s = str(root_r)
        if root_s in seen:
            continue
        seen.add(root_s)
        for name in IID_DICT_FILENAMES:
            direct = root_r / name
            if direct.exists():
                return direct
            for p in root_r.glob(f"*/{name}"):
                if p.exists():
                    return p
            for p in root_r.glob(f"*/*/{name}"):
                if p.exists():
                    return p
    return None


def configure_iid_dictionary(
    path: Path | None = None, search_roots: list[Path] | None = None
) -> str | None:
    global _IID_LOOKUP, _IID_DICTIONARY_SOURCE, _IID_LOOKUP_INITIALIZED
    resolved = discover_iid_dictionary_path(path, search_roots)
    _IID_LOOKUP_INITIALIZED = True
    if not resolved:
        _IID_LOOKUP = {}
        _IID_DICTIONARY_SOURCE = None
        return None
    _IID_LOOKUP = parse_iid_sid_item_dictionary(resolved)
    _IID_DICTIONARY_SOURCE = str(resolved)
    return _IID_DICTIONARY_SOURCE


def get_iid_dictionary_source() -> str | None:
    return _IID_DICTIONARY_SOURCE


def maybe_init_iid_dictionary() -> None:
    global _IID_LOOKUP_INITIALIZED
    if _IID_LOOKUP_INITIALIZED:
        return
    configure_iid_dictionary(None, None)


def resolve_ac_header_with_iid_dictionary(source_col: str) -> str | None:
    maybe_init_iid_dictionary()
    if not _IID_LOOKUP:
        return None

    candidates: list[str] = []
    candidates.extend(build_iid_lookup_keys(source_col))
    for code in EPA_CODE_RE.findall(source_col):
        candidates.extend(build_iid_lookup_keys(code))
    for sid in SID_CODE_RE.findall(source_col):
        candidates.extend(build_iid_lookup_keys(sid.replace("-", "_")))

    for key in candidates:
        iid = _IID_LOOKUP.get(key)
        if iid:
            return to_target_col_from_iid(iid)
    return None


def resolve_ac_legacy_header(source_col: str, target_set: set[str]) -> str | None:
    src_norm = normalize(source_col)
    if src_norm in AC_LEGACY_CASE_KEYS and "coCaseId" in target_set:
        return "coCaseId"
    if src_norm in AC_LEGACY_CASE_ALPHA_KEYS and "coCaseIdAlpha" in target_set:
        return "coCaseIdAlpha"

    m = AC_LEGACY_EPA_RE.match((source_col or "").strip().upper())
    if not m:
        return None
    num_raw = m.group(1)
    candidates: list[str] = []

    try:
        n_int = int(num_raw)
    except ValueError:
        n_int = -1

    if n_int >= 0:
        candidates.append(f"coE0I{n_int:03d}")
        candidates.append(f"coE0I{n_int}")

    candidates.append(f"coE0I{num_raw}")
    nz = num_raw.lstrip("0")
    if nz:
        candidates.append(f"coE0I{nz.zfill(3) if len(nz) < 3 else nz}")

    for width in (4, 5, 6):
        if len(num_raw) >= width:
            tail = num_raw[-width:]
            candidates.append(f"coE0I{tail}")
            tail_nz = tail.lstrip("0")
            if tail_nz:
                candidates.append(
                    f"coE0I{tail_nz.zfill(3) if len(tail_nz) < 3 else tail_nz}"
                )

    seen: set[str] = set()
    for cand in candidates:
        if cand in seen:
            continue
        seen.add(cand)
        if cand in target_set:
            return cand
    return None


def dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, raw in enumerate(headers, start=1):
        base = (raw or "").strip() or f"col_{i}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        out.append(base if count == 1 else f"{base}__{count}")
    return out


def parse_schema(sql_path: Path) -> dict[str, list[str]]:
    schema: dict[str, list[str]] = {}
    current_table: str | None = None
    for raw_line in sql_path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("/*"):
            continue
        create_match = CREATE_RE.match(line)
        if create_match:
            current_table = create_match.group(1)
            schema[current_table] = []
            continue
        if current_table is None:
            continue
        if line.startswith(")"):
            current_table = None
            continue
        if line.lower().startswith("constraint"):
            continue
        parts = line.rstrip(",").split()
        if len(parts) >= 2:
            schema[current_table].append(parts[0])
    return schema


def sniff_text_file(path: Path) -> tuple[str, str]:
    for enc in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = path.read_text(encoding=enc)
            first = text.splitlines()[0] if text.splitlines() else ""
            delim = ";" if first.count(";") > first.count(",") else ","
            return enc, delim
        except UnicodeDecodeError:
            continue
    return "latin1", ","


def cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    # Excel files saved from pandas often contain IEEE NaN floats.
    # Treat them as missing/empty to avoid writing literal "nan" everywhere.
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def is_null_like_text(value: Any) -> bool:
    text = "" if value is None else str(value).strip()
    if not text:
        return True
    return text.lower() in NULL_LIKE_TEXT


def sanitize_entity_id_fields(row: dict[str, Any]) -> None:
    for col in ENTITY_ID_COLS:
        if col not in row:
            continue
        if is_null_like_text(row.get(col)):
            row[col] = ""


def strip_trailing_empty(row: list[str]) -> list[str]:
    out = row[:]
    while out and out[-1] == "":
        out.pop()
    return out


def clean_sql_identifier(name: str) -> str:
    token = name.strip()
    if not token:
        return token
    if token.startswith("[") and token.endswith("]"):
        token = token[1:-1]
    elif token.startswith("`") and token.endswith("`"):
        token = token[1:-1]
    elif token.startswith('"') and token.endswith('"'):
        token = token[1:-1]
    return token.strip()


def clean_sql_table_name(name: str) -> str:
    parts = [clean_sql_identifier(p) for p in name.split(".") if p.strip()]
    if not parts:
        return clean_sql_identifier(name)
    return parts[-1]


def split_sql_csv(segment: str) -> list[str]:
    values: list[str] = []
    buf: list[str] = []
    in_single = False
    in_double = False
    depth = 0
    i = 0
    while i < len(segment):
        ch = segment[i]
        nxt = segment[i + 1] if i + 1 < len(segment) else ""
        if ch == "'" and not in_double:
            if in_single and nxt == "'":
                buf.append("''")
                i += 2
                continue
            in_single = not in_single
            buf.append(ch)
            i += 1
            continue
        if ch == '"' and not in_single:
            in_double = not in_double
            buf.append(ch)
            i += 1
            continue
        if not in_single and not in_double:
            if ch == "(":
                depth += 1
            elif ch == ")" and depth > 0:
                depth -= 1
            elif ch == "," and depth == 0:
                values.append("".join(buf).strip())
                buf = []
                i += 1
                continue
        buf.append(ch)
        i += 1
    tail = "".join(buf).strip()
    if tail:
        values.append(tail)
    return values


def extract_sql_value_tuples(values_blob: str) -> list[str]:
    tuples: list[str] = []
    depth = 0
    start = -1
    in_single = False
    in_double = False
    i = 0
    while i < len(values_blob):
        ch = values_blob[i]
        nxt = values_blob[i + 1] if i + 1 < len(values_blob) else ""
        if ch == "'" and not in_double:
            if in_single and nxt == "'":
                i += 2
                continue
            in_single = not in_single
            i += 1
            continue
        if ch == '"' and not in_single:
            in_double = not in_double
            i += 1
            continue
        if not in_single and not in_double:
            if ch == "(":
                if depth == 0:
                    start = i + 1
                depth += 1
            elif ch == ")" and depth > 0:
                depth -= 1
                if depth == 0 and start >= 0:
                    tuples.append(values_blob[start:i])
                    start = -1
        i += 1
    return tuples


def parse_sql_literal(raw: str) -> str:
    token = raw.strip()
    if not token:
        return ""
    up = token.upper()
    if up == "NULL" or up == "DEFAULT":
        return ""
    if token[:2].lower() == "n'" and token.endswith("'"):
        token = token[1:]
    if token.startswith("'") and token.endswith("'") and len(token) >= 2:
        return token[1:-1].replace("''", "'").strip()
    if token.startswith('"') and token.endswith('"') and len(token) >= 2:
        return token[1:-1].replace('""', '"').strip()
    return token.strip()


def parse_sql_insert_statements(text: str) -> list[dict[str, Any]]:
    statements: list[dict[str, Any]] = []
    for match in SQL_INSERT_RE.finditer(text):
        raw_table = match.group(1) or ""
        raw_columns = match.group(2) or ""
        raw_values = match.group(3) or ""
        table_name = clean_sql_table_name(raw_table)
        columns = [clean_sql_identifier(c) for c in split_sql_csv(raw_columns)] if raw_columns else []
        tuple_blobs = extract_sql_value_tuples(raw_values)
        rows: list[list[str]] = []
        for blob in tuple_blobs:
            values = [parse_sql_literal(v) for v in split_sql_csv(blob)]
            rows.append(values)
        statements.append(
            {
                "table_name": table_name,
                "columns": columns,
                "rows": rows,
            }
        )
    return statements


def rows_to_records(
    rows: list[list[str]], table: str, profile: dict[str, Any]
) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    if not rows:
        return [], [], profile

    row_width = max(len(r) for r in rows) if rows else 0
    normalized_rows = [r + [""] * (row_width - len(r)) for r in rows]

    if table == "tbImportAcData" and len(normalized_rows) >= 3:
        row2 = normalized_rows[1]
        if sum(1 for c in row2 if "E" in c and "_I_" in c) >= 10:
            headers = row2
            data_rows = normalized_rows[2:]
            profile["ac_mode"] = "row2_iids_header"
        else:
            headers = normalized_rows[0]
            data_rows = normalized_rows[1:]
            profile["ac_mode"] = "row1_header"
    else:
        headers = normalized_rows[0]
        data_rows = normalized_rows[1:]

    header_like = sum(1 for h in headers if re.search(r"[A-Za-z_]", h or "")) >= 1
    if not header_like:
        headers = [f"col_{i+1}" for i in range(len(normalized_rows[0]))]
        data_rows = normalized_rows
        profile["header_generated"] = True

    headers = dedupe_headers(headers)
    records = [{headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))} for row in data_rows]
    return headers, records, profile


def load_csv_records(path: Path, table: str) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    enc, delim = sniff_text_file(path)
    with path.open(encoding=enc, newline="") as f:
        rows = list(csv.reader(f, delimiter=delim))
    profile: dict[str, Any] = {"encoding": enc, "delimiter": delim, "format": "csv"}
    rows = [strip_trailing_empty([cell_to_str(c).strip() for c in row]) for row in rows]
    rows = [r for r in rows if any(c != "" for c in r)]
    return rows_to_records(rows, table, profile)


def load_xlsx_records(path: Path, table: str) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_rows = []
    for row in ws.iter_rows(values_only=True):
        normalized = strip_trailing_empty([cell_to_str(c).strip() for c in row])
        if any(c != "" for c in normalized):
            raw_rows.append(normalized)
    wb.close()

    profile: dict[str, Any] = {"format": "xlsx", "sheet": ws.title}
    if not raw_rows:
        return [], [], profile

    first = raw_rows[0]
    first_cell = first[0] if first else ""
    if first_cell and ("," in first_cell or ";" in first_cell):
        other_cells_empty = all(c == "" for c in first[1:])
        if other_cells_empty:
            delim = ";" if first_cell.count(";") > first_cell.count(",") else ","
            text_lines: list[str] = []
            for row in raw_rows:
                pieces = [c for c in row if c != ""]
                if not pieces:
                    continue
                text_lines.append(delim.join(pieces))
            csv_rows = [strip_trailing_empty(r) for r in csv.reader(text_lines, delimiter=delim)]
            csv_rows = [r for r in csv_rows if any(c != "" for c in r)]
            profile["xlsx_mode"] = "csv_in_cells"
            profile["delimiter"] = delim
            return rows_to_records(csv_rows, table, profile)

    profile["xlsx_mode"] = "tabular"
    return rows_to_records(raw_rows, table, profile)


def extract_nursing_page_record(page_text: str) -> dict[str, str]:
    flat = " ".join((page_text or "").split())
    m_date = PDF_DATE_RE.search(flat)
    m_shift = PDF_SHIFT_RE.search(flat)
    m_pat = PDF_PAT_RE.search(flat)
    m_case = PDF_CASE_RE.search(flat)
    m_ward = PDF_WARD_RE.search(flat)
    report_text = ""
    report_marker = re.search(r"\bReport\b", flat, re.IGNORECASE)
    if report_marker:
        report_text = flat[report_marker.end() :].strip()
    return {
        "case_id": m_case.group(1).strip() if m_case else "",
        "patient_id": m_pat.group(1).strip() if m_pat else "",
        "ward": m_ward.group(1).strip() if m_ward else "",
        "report_date": m_date.group(1).strip() if m_date else "",
        "shift": m_shift.group(1).strip() if m_shift else "",
        "nursing_note_free_text": report_text,
    }


def load_pdf_records(path: Path, table: str) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    profile: dict[str, Any] = {"format": "pdf"}
    with pdfplumber.open(path) as pdf:
        pages = pdf.pages
        profile["pages"] = len(pages)

        if table == "tbImportNursingDailyReportsData":
            rows: list[list[str]] = []
            for page in pages:
                record = extract_nursing_page_record(page.extract_text() or "")
                rows.append(
                    [
                        record["case_id"],
                        record["patient_id"],
                        record["ward"],
                        record["report_date"],
                        record["shift"],
                        record["nursing_note_free_text"],
                    ]
                )
            profile["pdf_mode"] = "page_text_nursing"
            return rows_to_records(
                [["case_id", "patient_id", "ward", "report_date", "shift", "nursing_note_free_text"]]
                + rows,
                table,
                profile,
            )

        table_rows: list[list[str]] = []
        for page in pages:
            for table_data in page.extract_tables() or []:
                for row in table_data:
                    if not row:
                        continue
                    normalized = strip_trailing_empty([cell_to_str(c).strip() for c in row])
                    if any(c != "" for c in normalized):
                        table_rows.append(normalized)
        if table_rows:
            profile["pdf_mode"] = "table_extract"
            return rows_to_records(table_rows, table, profile)

        text_rows: list[list[str]] = []
        for page in pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                clean = line.strip()
                if clean:
                    text_rows.append([clean])
        profile["pdf_mode"] = "line_text_fallback"
        return rows_to_records(text_rows, table, profile)


def load_sql_records(path: Path, table: str) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    enc, _ = sniff_text_file(path)
    text = path.read_text(encoding=enc)
    statements = parse_sql_insert_statements(text)
    profile: dict[str, Any] = {
        "format": "sql",
        "encoding": enc,
        "insert_statements": len(statements),
    }
    if not statements:
        return [], [], profile

    grouped: dict[str, list[dict[str, Any]]] = {}
    for stmt in statements:
        name = stmt["table_name"] or "unknown"
        grouped.setdefault(name, []).append(stmt)

    best_table = max(grouped.keys(), key=lambda t: sum(len(s["rows"]) for s in grouped[t]))
    selected = grouped[best_table]
    profile["sql_table"] = best_table
    profile["sql_tables_found"] = sorted(grouped.keys())

    headers: list[str] = []
    max_len_without_cols = 0
    for stmt in selected:
        cols = stmt["columns"]
        if cols:
            for col in cols:
                if col and col not in headers:
                    headers.append(col)
        else:
            for row in stmt["rows"]:
                if len(row) > max_len_without_cols:
                    max_len_without_cols = len(row)

    if not headers:
        headers = [f"col_{i+1}" for i in range(max_len_without_cols)]

    data_rows: list[list[str]] = []
    for stmt in selected:
        cols = stmt["columns"]
        for row in stmt["rows"]:
            out = [""] * len(headers)
            if cols:
                for idx, col in enumerate(cols):
                    if idx >= len(row):
                        continue
                    if col not in headers:
                        continue
                    out[headers.index(col)] = row[idx]
            else:
                for idx, val in enumerate(row[: len(headers)]):
                    out[idx] = val
            data_rows.append(out)

    return rows_to_records([headers] + data_rows, table, profile)


def load_records(path: Path, table: str) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    if path.suffix.lower() == ".xlsx":
        return load_xlsx_records(path, table)
    if path.suffix.lower() == ".pdf":
        return load_pdf_records(path, table)
    if path.suffix.lower() == ".sql":
        return load_sql_records(path, table)
    return load_csv_records(path, table)


def detect_table(
    path: Path,
    headers: list[str],
    profile: dict[str, Any] | None = None,
    debug: dict[str, Any] | None = None,
) -> str | None:
    if debug is not None:
        debug["input_file"] = str(path)
        debug["filename"] = path.name
        debug["headers_count"] = len(headers)
        debug["headers_sample"] = headers[:12]
        debug["profile_sql_table"] = (
            str(profile.get("sql_table")) if profile and profile.get("sql_table") else None
        )

    def matched(table: str, by: str, value: str | None = None) -> str:
        if debug is not None:
            debug["matched_table"] = table
            debug["matched_by"] = by
            debug["matched_value"] = value
        return table

    name = path.name.lower()
    if profile and profile.get("sql_table"):
        sql_name = normalize(str(profile["sql_table"]))
        target_by_norm = {normalize(t): t for t in TABLE_HINTS}
        if sql_name in target_by_norm:
            return matched(
                target_by_norm[sql_name],
                "sql_table_exact",
                str(profile["sql_table"]),
            )
        if "1hz" in sql_name:
            return matched("tbImportDevice1HzMotionData", "sql_table_keyword", "1hz")
        if "device" in sql_name:
            return matched("tbImportDeviceMotionData", "sql_table_keyword", "device")
        if "labs" in sql_name or "lab" in sql_name:
            return matched("tbImportLabsData", "sql_table_keyword", "labs")
        if "icd" in sql_name or "ops" in sql_name:
            return matched("tbImportIcd10Data", "sql_table_keyword", "icd/ops")
        if "med" in sql_name:
            return matched("tbImportMedicationInpatientData", "sql_table_keyword", "med")
        if "nursing" in sql_name or "report" in sql_name:
            return matched(
                "tbImportNursingDailyReportsData",
                "sql_table_keyword",
                "nursing/report",
            )
        if "acdata" in sql_name or "epaac" in sql_name:
            return matched("tbImportAcData", "sql_table_keyword", "acdata/epaac")

    if "1hz" in name:
        return matched("tbImportDevice1HzMotionData", "filename_keyword", "1hz")
    if "device" in name:
        return matched("tbImportDeviceMotionData", "filename_keyword", "device")
    if "labs" in name:
        return matched("tbImportLabsData", "filename_keyword", "labs")
    if "icd" in name or "ops" in name:
        return matched("tbImportIcd10Data", "filename_keyword", "icd/ops")
    if "med" in name:
        return matched("tbImportMedicationInpatientData", "filename_keyword", "med")
    if "nursing" in name or "report" in name:
        return matched("tbImportNursingDailyReportsData", "filename_keyword", "nursing/report")
    if "epaac" in name:
        return matched("tbImportAcData", "filename_keyword", "epaac")

    joined = " ".join(headers).lower()
    for table, hints in TABLE_HINTS.items():
        if all(h in joined for h in hints[:1]):
            return matched(table, "header_hint", hints[0])
    if debug is not None:
        debug["matched_table"] = None
        debug["matched_by"] = "none"
        debug["matched_value"] = None
    return None


DEFAULT_PROMPT_TEMPLATE = (
    "Map source columns to target SQL columns.\n"
    "Return JSON only.\n"
    "Keys: source columns exactly as given.\n"
    "Values: one target column from target list or null.\n"
    "Never invent target columns.\n\n"
    "source_columns={{SOURCE_COLUMNS}}\n"
    "target_columns={{TARGET_COLUMNS}}\n"
)


def build_prompt(
    source_columns: list[str], free_targets: list[str], prompt_template: str | None
) -> str:
    template = prompt_template or DEFAULT_PROMPT_TEMPLATE
    return (
        template.replace("{{SOURCE_COLUMNS}}", json.dumps(source_columns)).replace(
            "{{TARGET_COLUMNS}}", json.dumps(free_targets)
        )
    )


def run_llama_mapping(
    model: str,
    source_columns: list[str],
    free_targets: list[str],
    timeout_s: int,
    prompt_template: str | None,
) -> dict[str, str | None]:
    if not source_columns or not free_targets:
        return {}
    prompt = build_prompt(
        source_columns=source_columns,
        free_targets=free_targets,
        prompt_template=prompt_template,
    )
    try:
        result = subprocess.run(
            ["ollama", "run", model, prompt],
            check=False,
            capture_output=True,
            text=True,
            timeout=timeout_s,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return {}

    text = (result.stdout or "").strip()
    match = JSON_BLOCK_RE.search(text)
    if not match:
        return {}
    try:
        parsed = json.loads(match.group(0))
        out: dict[str, str | None] = {}
        for key, value in parsed.items():
            out[str(key)] = None if value is None else str(value)
        return out
    except json.JSONDecodeError:
        return {}


def build_mapping(
    table: str,
    headers: list[str],
    target_cols: list[str],
    use_llm: bool,
    model: str,
    timeout_s: int,
    prompt_template: str | None,
) -> dict[str, str]:
    target_set = set(target_cols)
    mapping: dict[str, str] = {}
    alias_table = ALIASES.get(table, {})
    common_case_keys = {
        "caseid",
        "case_id",
        "fallid",
        "fall_id",
        "fall",
        "cas",
        "idcas",
        "id_case",
    }
    common_patient_keys = {
        "pid",
        "patientid",
        "patient_id",
        "patid",
        "pat_id",
        "idpat",
        "id_pat",
        "pat",
    }
    common_encounter_keys = {
        "encounterid",
        "encounter_id",
        "encid",
        "enc_id",
        "idenc",
        "id_enc",
    }

    # 1) Exact normalized match: source -> co<source>
    target_normalized = {normalize(c.removeprefix("co")): c for c in target_cols}
    for source_col in headers:
        src_norm = normalize(source_col)
        if src_norm in alias_table:
            dest = alias_table[src_norm]
            if dest in target_set:
                mapping[source_col] = dest
                continue
        if "coCaseId" in target_set and (
            src_norm in common_case_keys
            or src_norm.startswith("fall")
            or src_norm.startswith("case")
        ):
            mapping[source_col] = "coCaseId"
            continue
        if "coPatient_id" in target_set and (
            src_norm in common_patient_keys
            or src_norm.startswith("patient")
        ):
            mapping[source_col] = "coPatient_id"
            continue
        if "coEncounter_id" in target_set and (
            src_norm in common_encounter_keys
            or src_norm.startswith("encounter")
        ):
            mapping[source_col] = "coEncounter_id"
            continue
        if table == "tbImportAcData":
            legacy_dest = resolve_ac_legacy_header(source_col, target_set)
            if legacy_dest:
                mapping[source_col] = legacy_dest
                continue
            codes = EPA_CODE_RE.findall(source_col)
            for code in codes:
                candidate = "co" + code.replace("_", "").upper()
                if candidate in target_set:
                    mapping[source_col] = candidate
                    break
            if source_col in mapping:
                continue
            dict_dest = resolve_ac_header_with_iid_dictionary(source_col)
            if dict_dest and dict_dest in target_set:
                mapping[source_col] = dict_dest
                continue
        dest = target_normalized.get(src_norm)
        if dest:
            mapping[source_col] = dest

    # 2) Llama fallback for unresolved headers
    unresolved = [h for h in headers if h not in mapping]
    free_targets = [c for c in target_cols if c not in mapping.values()]
    if use_llm and unresolved and free_targets:
        suggestions = run_llama_mapping(
            model, unresolved, free_targets, timeout_s, prompt_template
        )
        for src, dest in suggestions.items():
            if src in unresolved and dest in target_set:
                mapping[src] = dest

    # Keep only one source per target column (first source column wins).
    # This avoids unstable overwrites when many headers map to the same destination.
    deduped: dict[str, str] = {}
    used_targets: set[str] = set()
    for source_col in headers:
        dest = mapping.get(source_col)
        if not dest or dest in used_targets:
            continue
        deduped[source_col] = dest
        used_targets.add(dest)
    mapping = deduped

    return mapping


def detect_ac_sid_columns(headers: list[str]) -> tuple[str | None, str | None]:
    sid_col: str | None = None
    val_col: str | None = None
    for col in headers:
        n = normalize(col)
        if sid_col is None and n in AC_SID_COL_KEYS:
            sid_col = col
        if val_col is None and n in AC_SID_VALUE_COL_KEYS:
            val_col = col
    return sid_col, val_col


def standardize_records(
    records: list[dict[str, Any]],
    target_cols: list[str],
    mapping: dict[str, str],
    table: str | None = None,
    headers: list[str] | None = None,
) -> list[dict[str, Any]]:
    out_rows: list[dict[str, Any]] = []
    sid_col = None
    sid_value_col = None
    if table == "tbImportAcData" and headers:
        sid_col, sid_value_col = detect_ac_sid_columns(headers)
    target_set = set(target_cols)

    for rec in records:
        out = {c: "" for c in target_cols}
        for src, dest in mapping.items():
            if src not in rec:
                continue
            if src == sid_col or src == sid_value_col:
                continue
            val = rec[src]
            if val is None:
                continue
            val = str(val).strip()
            if not val:
                continue
            out[dest] = val

        # AC long format (SID + SID_value): dynamic row-wise mapping
        # e.g., SID=00_14 and SID_value=70 -> coE0I014=70
        if table == "tbImportAcData" and sid_col and sid_value_col:
            sid_raw = str(rec.get(sid_col, "") or "").strip()
            val_raw = str(rec.get(sid_value_col, "") or "").strip()
            if sid_raw and val_raw:
                dyn_dest = resolve_ac_header_with_iid_dictionary(sid_raw)
                if dyn_dest and dyn_dest in target_set:
                    out[dyn_dest] = val_raw

        # Keep null-like normalization limited to entity identifiers only.
        # All non-ID variables preserve their raw mapped values.
        sanitize_entity_id_fields(out)

        out_rows.append(out)
    return out_rows


def write_csv(path: Path, rows: list[dict[str, Any]], cols: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--schema-sql", required=True, type=Path)
    parser.add_argument("--input-dir", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--model", default="llama3.2:latest")
    parser.add_argument("--llm-timeout", default=45, type=int)
    parser.add_argument("--prompt-template", type=Path)
    parser.add_argument("--iid-dictionary", type=Path)
    parser.add_argument("--no-llm", action="store_true")
    args = parser.parse_args()

    schema = parse_schema(args.schema_sql)
    in_dir = args.input_dir
    out_dir = args.output_dir
    report: list[dict[str, Any]] = []
    prompt_template = None
    if args.prompt_template:
        prompt_template = args.prompt_template.read_text(encoding="utf-8")
    iid_source = configure_iid_dictionary(
        path=args.iid_dictionary,
        search_roots=[in_dir, in_dir.parent, Path.cwd(), Path(__file__).resolve().parents[1]],
    )

    files = sorted(
        [
            p
            for p in in_dir.rglob("*")
            if p.suffix.lower() in {".csv", ".xlsx", ".pdf", ".sql"}
        ]
    )
    for path in files:
        try:
            headers_probe, _, profile_probe = load_records(path, "tbImportAcData")
            routing_debug: dict[str, Any] = {}
            table = detect_table(path, headers_probe, profile_probe, debug=routing_debug)
            if not table or table not in schema:
                log_routing_debug(
                    {
                        "file": str(path),
                        "status": "skipped",
                        "reason": "table_not_detected_or_not_in_schema",
                        "routing_debug": routing_debug,
                    }
                )
                report.append(
                    {
                        "file": str(path),
                        "status": "skipped",
                        "reason": "table_not_detected_or_not_in_schema",
                        "routing_debug": routing_debug,
                    }
                )
                continue

            headers, records, profile = load_records(path, table)
            target_cols = [c for c in schema[table] if c != "coId"]
            routing_debug["target_cols_count"] = len(target_cols)
            routing_debug["target_cols_sample"] = target_cols[:12]
            profile["routing_debug"] = routing_debug
            log_routing_debug(
                {
                    "file": str(path),
                    "status": "ok",
                    "table": table,
                    "routing_debug": routing_debug,
                }
            )
            mapping = build_mapping(
                table=table,
                headers=headers,
                target_cols=target_cols,
                use_llm=not args.no_llm,
                model=args.model,
                timeout_s=args.llm_timeout,
                prompt_template=prompt_template,
            )
            rows = standardize_records(
                records=records,
                target_cols=target_cols,
                mapping=mapping,
                table=table,
                headers=headers,
            )

            ext = path.suffix.lower().lstrip(".")
            output_name = f"{path.stem}__{ext}__{table}.csv"
            output_path = out_dir / output_name
            write_csv(output_path, rows, target_cols)

            report.append(
                {
                    "file": str(path),
                    "table": table,
                    "status": "ok",
                    "rows_in": len(records),
                    "rows_out": len(rows),
                    "headers_in": len(headers),
                    "targets_total": len(target_cols),
                    "mapped_columns": len(mapping),
                    "unmapped_headers": [h for h in headers if h not in mapping],
                    "profile": profile,
                    "iid_dictionary": iid_source,
                    "routing_debug": routing_debug,
                }
            )
        except Exception as exc:
            report.append(
                {
                    "file": str(path),
                    "status": "error",
                    "reason": str(exc),
                }
            )
            continue

    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "standardization_report.json").write_text(
        json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
